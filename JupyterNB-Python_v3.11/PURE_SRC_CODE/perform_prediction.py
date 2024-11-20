import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import argparse
from scipy.stats import shapiro, kstest, norm, probplot, chi2_contingency
from sklearn.model_selection import cross_val_score
from sklearn.metrics import (accuracy_score, log_loss, mean_squared_error, confusion_matrix,
                             precision_score, recall_score, auc, roc_curve, roc_auc_score,
                             f1_score, PrecisionRecallDisplay, RocCurveDisplay,
                             ConfusionMatrixDisplay, mean_absolute_error)
from sklearn.metrics import classification_report
from sklearn.calibration import calibration_curve
import configparser
import joblib
from sklearn.preprocessing import label_binarize
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Function to load folder paths from config
def load_paths(config_file):
    config = configparser.ConfigParser()
    config.read(config_file)
    paths = {
        "input_folder_predict": config["Paths"].get("input_folder_predict", "../OTH_DATA/cleaned_data"),
        "model_folder_predict": config["Paths"].get("model_folder_predict", "../ML_DATA/model_outputs"),
        "output_folder_predict": config["Paths"].get("output_folder_predict", "../ML_DATA/predict_outputs")
    }
    return paths

def select_testing_file(input_folder):
    files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]
    if not files:
        print("No CSV files found in the testing_data folder.")
        return None
    print("Available files for testing:")
    for idx, file in enumerate(files, 1):
        print(f"{idx}: {file}")
    choice = int(input("Select the file number to test the model on: ")) - 1
    return os.path.join(input_folder, files[choice])

def load_data(file_path):
    data = pd.read_csv(file_path)
    target_col = "Obesity_Level"
    
    if target_col in data.columns:
        # Target column is present
        X_test = data.drop(columns=[target_col])
        y_test = data[target_col]
        print(f"Target column '{target_col}' found. Data loaded with features and target.")
    else:
        # Target column is absent
        X_test = data  # Use all columns as features
        y_test = None  # No target available
        print(f"Target column '{target_col}' not found. Data loaded with features only.")
    
    return X_test, y_test


def load_data(file_path):
    data = pd.read_csv(file_path)
    target_col = "Obesity_Level"
    
    if target_col in data.columns:
        # Target column is present
        X_test = data.drop(columns=[target_col])
        y_test = data[target_col]
        print(f"Target column '{target_col}' found. Data loaded with features and target.")
    else:
        # Target column is absent
        X_test = data  # Use all columns as features
        y_test = None  # No target available
        print(f"Target column '{target_col}' not found. Data loaded with features only.")
    
    return X_test, y_test


def load_models(model_folder):
    models = {}
    for model_file in os.listdir(model_folder):
        if model_file.endswith('.pkl'):
            model_path = os.path.join(model_folder, model_file)
            model_name = model_file.split('_model.pkl')[0]
            models[model_name] = joblib.load(model_path)
    return models

def preprocess_test_data(X_test, training_features):
    for col in training_features:
        if col not in X_test.columns:
            X_test[col] = 0
    X_test = X_test[training_features]
    return X_test

def output_predictions_with_formatting(X_test, y_test, y_pred, output_path):
    combined_mappings = {
        "MTRANS": {
            0: "Automobile",
            1: "Bike",
            2: "Motorbike",
            3: "Public_Transportation",
            4: "Walking"
        },
        "Obesity_Level": {
            0: "Insufficient_Weight",
            1: "Normal_Weight",
            2: "Obesity_Type_I",
            3: "Obesity_Type_II",
            4: "Obesity_Type_III",
            5: "Overweight_Level_I",
            6: "Overweight_Level_II"
        },
        "Gender": {0: "Female", 1: "Male"},
        "fam_hist_over-wt": {0: "no", 1: "yes"},
        "FAVC": {0: "no", 1: "yes"},
        "CAEC": {0: "Always", 1: "Frequently", 2: "Sometimes", 3: "no"},
        "SMOKE": {0: "no", 1: "yes"},
        "SCC": {0: "no", 1: "yes"},
        "CALC": {0: "Frequently", 1: "Sometimes", 2: "no"}
    }

    # Decode columns in X_test using the combined mappings
    for col, mapping in combined_mappings.items():
        if col in X_test.columns:
            X_test[col] = X_test[col].map(mapping)

    # Add the predicted column
    X_test["Predicted_Obesity_Level"] = pd.Series(y_pred).map(combined_mappings["Obesity_Level"])

    # Add the actual column only if y_test is not None
    if y_test is not None:
        X_test["Actual_Obesity_Level"] = y_test.map(combined_mappings["Obesity_Level"])

    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Add header row
    headers = list(X_test.columns)
    ws.append(headers)

    # Define styles for correct and incorrect predictions
    correct_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    incorrect_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

    # Add data rows with conditional formatting
    for i, row in X_test.iterrows():
        ws.append(row.tolist())
        # Apply formatting to the last column (Predicted_Obesity_Level)
        if y_test is not None:
            predicted_cell = ws.cell(row=i + 2, column=len(headers))  # Adjust for header row
            if row["Actual_Obesity_Level"] == row["Predicted_Obesity_Level"]:
                predicted_cell.fill = correct_fill
            else:
                predicted_cell.fill = incorrect_fill

    # Save the Excel file
    wb.save(output_path)
    print(f"Predictions with formatting saved to {output_path}")

def plot_confusion_matrix(cm, labels, output_path):
    # Increase the figure size
    plt.figure(figsize=(10, 8))  # Adjust size as needed
    
    # Create the confusion matrix plot
    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=labels)
    disp.plot(cmap='viridis', xticks_rotation=90, values_format='d')  # Rotate x-axis labels and set format

    # Adjust font size
    plt.title("Confusion Matrix", fontsize=16)
    plt.xlabel("Predicted label", fontsize=14)
    plt.ylabel("True label", fontsize=14)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)

    # Save and close the plot
    plt.tight_layout()  # Ensures everything fits without overlap
    plt.savefig(output_path)
    plt.close()
    print(f"Confusion Matrix saved to {output_path}")

def evaluate_model(model, X_test, y_test, output_dir, feature_names):
    y_pred = model.predict(X_test)
    y_prob = model.predict_proba(X_test) if hasattr(model, "predict_proba") else None

    # Define mappings
    combined_mappings = {
        "Obesity_Level": {
            0: "Insufficient_Weight",
            1: "Normal_Weight",
            2: "Obesity_Type_I",
            3: "Obesity_Type_II",
            4: "Obesity_Type_III",
            5: "Overweight_Level_I",
            6: "Overweight_Level_II"
        }
    }

    # Decode labels to string form
    if y_test is not None:
        y_test_decoded = y_test.map(combined_mappings["Obesity_Level"])
        y_pred_decoded = pd.Series(y_pred).map(combined_mappings["Obesity_Level"])
    else:
        y_test_decoded = None
        y_pred_decoded = pd.Series(y_pred).map(combined_mappings["Obesity_Level"])

    # Call the feature importance plotting function
    plot_feature_importances(model, output_dir)

    # Generate Excel output with conditional formatting
    output_predictions_with_formatting(
        X_test.copy(),  # Pass a copy to avoid modifying the original
        y_test,
        y_pred,
        os.path.join(output_dir, "predictions_with_formatting.xlsx")
    )

    # Handle cases where the target column (y_test) is not available
    if y_test is None:
        print("Target column is not present in the dataset. Skipping evaluation metrics calculation.")
        return {
            'accuracy': None,
            'precision': None,
            'recall': None,
            'f1_score': None,
            'roc_auc': None,
            'mse': None,
            'mae': None,
            'rmse': None,
            'confusion_matrix': None,
            'classification_report': "Target column is missing; no report generated."
        }

    # Validate test dataset classes
    unique_classes = np.unique(y_test)
    print(f"Unique classes in test set: {unique_classes}")

    # Check if the target is binary or multiclass
    is_multiclass = len(unique_classes) > 2

    # Metrics
    mse = mean_squared_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)
    rmse = np.sqrt(mse)

    # Create directory for the model if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # ROC Curve and AUC for binary or multiclass
    if y_prob is not None:
        if is_multiclass:
            print("Handling multiclass ROC curves...")
            y_test_bin = label_binarize(y_test, classes=unique_classes)
            n_classes = y_test_bin.shape[1]

            # Compute ROC curve and AUC for each class
            plt.figure()
            for i in range(n_classes):
                fpr, tpr, _ = roc_curve(y_test_bin[:, i], y_prob[:, i])
                roc_auc = auc(fpr, tpr)
                plt.plot(fpr, tpr, lw=2, label=f'Class {unique_classes[i]} (AUC = {roc_auc:.2f})')

            plt.plot([0, 1], [0, 1], color='gray', linestyle='--')
            plt.title("Multiclass ROC Curve")
            plt.xlabel("False Positive Rate")
            plt.ylabel("True Positive Rate")
            plt.legend(loc="best")
            plt.savefig(os.path.join(output_dir, "multiclass_roc_curve.png"))
            plt.close()
        else:
            print("Plotting binary ROC curve...")
            RocCurveDisplay.from_predictions(y_test, y_prob[:, 1])
            plt.title("ROC Curve")
            plt.savefig(os.path.join(output_dir, "roc_curve.png"))
            plt.close()

    # Confusion Matrix
    cm = confusion_matrix(y_test_decoded, y_pred_decoded, labels=list(combined_mappings["Obesity_Level"].values()))
    output_cm_path = os.path.join(output_dir, "confusion_matrix.png")
    plot_confusion_matrix(cm, list(combined_mappings["Obesity_Level"].values()), output_cm_path)

    metrics = {
        'accuracy': accuracy_score(y_test, y_pred),
        'precision': precision_score(y_test, y_pred, average="macro", zero_division=0),
        'recall': recall_score(y_test, y_pred, average="macro", zero_division=0),
        'f1_score': f1_score(y_test, y_pred, average="macro", zero_division=0),
        'roc_auc': None if y_prob is None else (
            roc_auc_score(y_test, y_prob, average="macro", multi_class="ovr") if is_multiclass else roc_auc_score(y_test, y_prob[:, 1])
        ),
        'mse': mse,
        'mae': mae,
        'rmse': rmse,
        'confusion_matrix': cm.tolist(),
        # 'classification_report': classification_report(y_test, y_pred, zero_division=0)
        'classification_report': classification_report(y_test_decoded, y_pred_decoded, zero_division=0)
    }

    return metrics

def plot_feature_importances(model, output_dir):
    # Access the underlying estimator if model is a pipeline
    if hasattr(model, 'named_steps'):
        preprocessor = model.named_steps['preprocessor']
        estimator = model.named_steps['model']
        feature_names = model.feature_names
    else:
        estimator = model
        feature_names = None

    # Check for feature_importances_
    if hasattr(estimator, "feature_importances_"):
        importances = estimator.feature_importances_
        print("Using feature_importances_ attribute.")
    else:
        print("Feature importances are not available for this model.")
        return

    # Use the stored feature names if available
    if feature_names is not None:
        feature_importances = pd.DataFrame({
            "Feature": feature_names,
            "Importance": importances
        }).sort_values(by="Importance", ascending=False)
    else:
        feature_importances = pd.DataFrame({
            "Feature": [f"Feature {i}" for i in range(len(importances))],
            "Importance": importances
        }).sort_values(by="Importance", ascending=False)

    # Print feature importances
    print("Feature Importances:")
    print(feature_importances)

    # Plot feature importances
    plt.figure(figsize=(10, 6))
    plt.barh(feature_importances["Feature"], feature_importances["Importance"], color="skyblue")
    plt.xlabel("Importance")
    plt.ylabel("Feature")
    plt.title("Feature Importances")
    plt.gca().invert_yaxis()  # Reverse the order for readability
    plt.tight_layout()

    # Save the plot
    output_path = os.path.join(output_dir, "feature_importances.png")
    plt.savefig(output_path)
    print(f"Feature importances plot saved to {output_path}")
    plt.close()

# Function to list and select models from the model folder
def select_models(model_folder):
    model_files = [f for f in os.listdir(model_folder) if f.endswith('.pkl')]
    if not model_files:
        print("No model files found in the specified model folder.")
        return {}
    
    print("Available models for prediction:")
    for idx, model_file in enumerate(model_files, 1):
        print(f"{idx}. {model_file}")
    
    selection = input("Enter the model numbers to use for prediction (comma-separated) or 'all' to use all models: ")
    if selection.lower() == 'all':
        selected_files = model_files
    else:
        selected_indices = [int(i.strip()) - 1 for i in selection.split(",")]
        selected_files = [model_files[i] for i in selected_indices]
    
    models = {}
    for model_file in selected_files:
        model_path = os.path.join(model_folder, model_file)
        model_name = model_file.split('_model.pkl')[0]
        models[model_name] = joblib.load(model_path)
    
    return models

def main(config_file="config.txt"):
    visialize_output_folder = "../ML_DATA/visualize_output"
    # Load paths from config
    paths = load_paths(config_file)

    # Select and load test file
    test_file = select_testing_file(paths["input_folder_predict"])
    if not test_file:
        print("No valid file selected. Exiting.")
        return

    X_test, y_test = load_data(test_file)

    # Select models to use for prediction
    models = select_models(paths["model_folder_predict"])
    if not models:
        print("No models selected for prediction. Exiting.")
        return

    
    report = []
    for model_name, model in models.items():
        # Extract feature names from the model
        try:
            feature_names = model.feature_names  # Assuming you saved feature_names in training script
        except AttributeError:
            print(f"Could not extract feature names from model {model_name}.")
            continue
        model_output_dir = os.path.join(visialize_output_folder, model_name)
        # metrics = evaluate_model(model, X_test, y_test, model_output_dir)
        metrics = evaluate_model(model, X_test, y_test, model_output_dir, feature_names)
        report.append((model_name, metrics))
        # print(f"Model: {model_name} - Metrics: {metrics}")

        # Format and print the metrics
        print(f"\n{'='*40}")
        print(f"Model: {model_name}")
        print(f"{'-'*40}")
        if metrics['classification_report'] != "Target column is missing; no report generated.":
            print("Classification Report:")
            print(metrics['classification_report'])
        else:
            print("Classification Report: Not available (Target column missing).")

        if metrics['accuracy'] is not None:
            print(f"Accuracy       : {metrics['accuracy']:.4f}")
        else:
            print("Accuracy       : Not available (Target column missing).")

        if metrics['roc_auc'] is not None:
            print(f"ROC AUC        : {metrics['roc_auc']:.4f}")
        else:
            print("ROC AUC        : Not available (e.g., no probabilities available or target missing).")

        if metrics['mse'] is not None:
            print(f"MSE            : {metrics['mse']:.4f}")
        else:
            print("MSE            : Not available (Target column missing).")

        if metrics['mae'] is not None:
            print(f"MAE            : {metrics['mae']:.4f}")
        else:
            print("MAE            : Not available (Target column missing).")

        if metrics['rmse'] is not None:
            print(f"RMSE           : {metrics['rmse']:.4f}")
        else:
            print("RMSE           : Not available (Target column missing).")

        if metrics['confusion_matrix'] is not None:
            print("Confusion Matrix:")
            print(metrics['confusion_matrix'])
        else:
            print("Confusion Matrix: Not available (Target column missing).")

        print(f"{'='*40}\n")

    report_path = os.path.join(visialize_output_folder, 'prediction_report.csv')
    report_df = pd.DataFrame([{**{'Model': m}, **metrics} for m, metrics in report])
    report_df.to_csv(report_path, index=False)
    print(f"Prediction report with confusion matrices saved to {report_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run model predictions and evaluations.")
    parser.add_argument('--config_file', type=str, default="config.txt", help="Path to the configuration file.")
    
    args = parser.parse_args()
    main(args.config_file)
