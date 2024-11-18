import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import argparse
from scipy.stats import shapiro, kstest, norm, probplot, chi2_contingency
from sklearn.model_selection import cross_val_score
from sklearn.metrics import (accuracy_score, log_loss, mean_squared_error, confusion_matrix,
                             precision_score, recall_score, auc, roc_curve, roc_auc_score,
                             f1_score, PrecisionRecallDisplay, RocCurveDisplay,
                             ConfusionMatrixDisplay, mean_absolute_error)
from sklearn.metrics import classification_report
from sklearn.calibration import calibration_curve
import configparser
import joblib
from sklearn.preprocessing import label_binarize
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Function to load folder paths from config
def load_paths(config_file):
    config = configparser.ConfigParser()
    config.read(config_file)
    paths = {
        "input_folder_predict": config["Paths"].get("input_folder_predict", "../OTH_DATA/cleaned_data"),
        "model_folder_predict": config["Paths"].get("model_folder_predict", "../ML_DATA/model_outputs"),
        "output_folder_predict": config["Paths"].get("output_folder_predict", "../ML_DATA/predict_outputs")
    }
    return paths

def select_testing_file(input_folder):
    files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]
    if not files:
        print("No CSV files found in the testing_data folder.")
        return None
    print("Available files for testing:")
    for idx, file in enumerate(files, 1):
        print(f"{idx}: {file}")
    choice = int(input("Select the file number to test the model on: ")) - 1
    return os.path.join(input_folder, files[choice])

def load_data(file_path):
    data = pd.read_csv(file_path)
    target_col = "Obesity_Level"
    X_test = data.drop(columns=[target_col])  
    y_test = data[target_col]  
    return X_test, y_test

def load_models(model_folder):
    models = {}
    for model_file in os.listdir(model_folder):
        if model_file.endswith('.pkl'):
            model_path = os.path.join(model_folder, model_file)
            model_name = model_file.split('_model.pkl')[0]
            models[model_name] = joblib.load(model_path)
    return models

def preprocess_test_data(X_test, training_features):
    for col in training_features:
        if col not in X_test.columns:
            X_test[col] = 0
    X_test = X_test[training_features]
    return X_test

def output_predictions_with_formatting(X_test, y_test, y_pred, output_path):
    # Add actual and predicted columns to the dataset
    X_test["Actual_Obesity_Level"] = y_test
    X_test["Predicted_Obesity_Level"] = y_pred

    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Add header row
    headers = list(X_test.columns)
    ws.append(headers)

    # Define styles for correct and incorrect predictions
    correct_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    incorrect_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

    # Add data rows with conditional formatting
    for i, row in X_test.iterrows():
        ws.append(row.tolist())
        # Apply formatting to the last column (Predicted_Obesity_Level)
        predicted_cell = ws.cell(row=i + 2, column=len(headers))  # Adjust for header row
        if row["Actual_Obesity_Level"] == row["Predicted_Obesity_Level"]:
            predicted_cell.fill = correct_fill
        else:
            predicted_cell.fill = incorrect_fill

    # Save the Excel file
    wb.save(output_path)
    print(f"Predictions with formatting saved to {output_path}")


def evaluate_model(model, X_test, y_test, output_dir, feature_names):
    y_pred = model.predict(X_test)
    y_prob = model.predict_proba(X_test) if hasattr(model, "predict_proba") else None

    # Generate Excel output with conditional formatting
    output_predictions_with_formatting(
        X_test.copy(),  # Pass a copy to avoid modifying the original
        y_test,
        y_pred,
        os.path.join(output_dir, "predictions_with_formatting.xlsx")
    )

    # Validate test dataset classes
    unique_classes = np.unique(y_test)
    print(f"Unique classes in test set: {unique_classes}")

    # Check if the target is binary or multiclass
    is_multiclass = len(unique_classes) > 2

    # Metrics
    mse = mean_squared_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)
    rmse = np.sqrt(mse)

    # Create directory for the model if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # ROC Curve and AUC for binary or multiclass
    if y_prob is not None:
        if is_multiclass:
            print("Handling multiclass ROC curves...")
            y_test_bin = label_binarize(y_test, classes=unique_classes)
            n_classes = y_test_bin.shape[1]

            # Compute ROC curve and AUC for each class
            plt.figure()
            for i in range(n_classes):
                fpr, tpr, _ = roc_curve(y_test_bin[:, i], y_prob[:, i])
                roc_auc = auc(fpr, tpr)
                plt.plot(fpr, tpr, lw=2, label=f'Class {unique_classes[i]} (AUC = {roc_auc:.2f})')

            plt.plot([0, 1], [0, 1], color='gray', linestyle='--')
            plt.title("Multiclass ROC Curve")
            plt.xlabel("False Positive Rate")
            plt.ylabel("True Positive Rate")
            plt.legend(loc="best")
            plt.savefig(os.path.join(output_dir, "multiclass_roc_curve.png"))
            plt.close()
        else:
            print("Plotting binary ROC curve...")
            RocCurveDisplay.from_predictions(y_test, y_prob[:, 1])
            plt.title("ROC Curve")
            plt.savefig(os.path.join(output_dir, "roc_curve.png"))
            plt.close()

    # Confusion Matrix
    cm = confusion_matrix(y_test, y_pred)
    ConfusionMatrixDisplay(cm).plot()
    plt.title("Confusion Matrix")
    plt.savefig(os.path.join(output_dir, "confusion_matrix.png"))
    plt.close()

    # Metrics dictionary
    # metrics = {
    #     'accuracy': accuracy_score(y_test, y_pred),
    #     'precision': precision_score(y_test, y_pred, average="macro", zero_division=0),
    #     'recall': recall_score(y_test, y_pred, average="macro", zero_division=0),
    #     'f1_score': f1_score(y_test, y_pred, average="macro", zero_division=0),
    #     'roc_auc': None if y_prob is None else (
    #         roc_auc_score(y_test, y_prob, average="macro", multi_class="ovr") if is_multiclass else roc_auc_score(y_test, y_prob[:, 1])
    #     ),
    #     'mse': mse,
    #     'mae': mae,
    #     'rmse': rmse,
    #     'confusion_matrix': cm.tolist(),
    #     'classification_report': classification_report(y_test, y_pred, zero_division=0)
    # }
    metrics = {
        'accuracy': accuracy_score(y_test, y_pred),
        'precision': precision_score(y_test, y_pred, average="macro", zero_division=0),
        'recall': recall_score(y_test, y_pred, average="macro", zero_division=0),
        'f1_score': f1_score(y_test, y_pred, average="macro", zero_division=0),
        'roc_auc': None if y_prob is None else (
            roc_auc_score(y_test, y_prob, average="macro", multi_class="ovr") if is_multiclass else roc_auc_score(y_test, y_prob[:, 1])
        ),
        'mse': mse,
        'mae': mae,
        'rmse': rmse,
        'confusion_matrix': cm.tolist(),
        'classification_report': classification_report(y_test, y_pred, zero_division=0)
    }

    return metrics

# def evaluate_model(model, X_test, y_test, output_dir, feature_names):
    y_pred = model.predict(X_test)
    y_prob = model.predict_proba(X_test)[:, 1] if hasattr(model, "predict_proba") else None

    # Calculate MSE, MAE, and RMSE
    mse = mean_squared_error(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)
    rmse = np.sqrt(mse)

    # Create directory for the model if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Confusion Matrix
    cm = confusion_matrix(y_test, y_pred)
    ConfusionMatrixDisplay(cm).plot()
    plt.title("Confusion Matrix")
    plt.savefig(os.path.join(output_dir, "confusion_matrix.png"))
    plt.close()

    # ROC Curve and AUC
    if y_prob is not None:
        RocCurveDisplay.from_predictions(y_test, y_prob)
        plt.title("ROC Curve")
        plt.savefig(os.path.join(output_dir, "roc_curve.png"))
        plt.close()

    # Precision-Recall Curve
    if y_prob is not None:
        PrecisionRecallDisplay.from_predictions(y_test, y_prob)
        plt.title("Precision-Recall Curve")
        plt.savefig(os.path.join(output_dir, "precision_recall_curve.png"))
        plt.close()

    # Calibration Curve
    if y_prob is not None:
        prob_true, prob_pred = calibration_curve(y_test, y_prob, n_bins=10)
        plt.plot(prob_pred, prob_true, marker='o')
        plt.plot([0, 1], [0, 1], linestyle="--", color='gray')
        plt.title("Calibration Curve")
        plt.xlabel("Mean Predicted Probability")
        plt.ylabel("Fraction of Positives")
        plt.savefig(os.path.join(output_dir, "calibration_curve.png"))
        plt.close()

    # Identify missing features, EDIT
    missing_features = [feature for feature in feature_names if feature not in X_test.columns]
    if missing_features:
        # print("\nThe following features were missing in the test dataset:")
        for feature in missing_features:
            print(f"- {feature}")
        # print("They have been filled with zeros in the test data.")
    # Access the final estimator
    final_estimator = model.named_steps['model']
    # Feature Importance or Coefficients (for applicable models)
    if hasattr(final_estimator, 'feature_importances_'):
        importances = final_estimator.feature_importances_
        indices = np.argsort(importances)[::-1]
        print("Feature importances:")
        for i in indices:
            feature_name = feature_names[i]
            importance = importances[i]
            # status = " (missing in test data)" if feature_name in missing_features else ""
            status = "" if feature_name in missing_features else ""
            print(f"{feature_name}: {importance}{status}")

        # Plot feature importances
        plt.figure()
        plt.title("Feature Importances")
        plt.bar(range(len(importances)), importances[indices], align="center")
        plt.xticks(range(len(importances)), [feature_names[i] for i in indices], rotation=90)
        plt.tight_layout()  # Adjust layout to prevent label cutoff
        feature_importance_path = os.path.join(output_dir, "feature_importance.png")
        plt.savefig(feature_importance_path)
        plt.close()
        print(f"Feature importance plot saved for {model} at {feature_importance_path}")
    elif hasattr(final_estimator, 'coef_'):
        importances = np.abs(final_estimator.coef_).flatten()
        indices = np.argsort(importances)[::-1]
        print("Feature coefficients (importance):")
        for i in indices:
            feature_name = feature_names[i]
            importance = importances[i]
            # status = " (missing in test data)" if feature_name in missing_features else ""
            status = "" if feature_name in missing_features else ""
            print(f"{feature_name}: {importance}{status}")

        # Plot feature coefficients
        plt.figure()
        plt.title("Feature Coefficients (Importance)")
        plt.bar(range(len(importances)), importances[indices], align="center")
        plt.xticks(range(len(importances)), [feature_names[i] for i in indices], rotation=90)
        plt.tight_layout()
        feature_coefficients_path = os.path.join(output_dir, "feature_coefficients.png")
        plt.savefig(feature_coefficients_path)
        plt.close()
        print(f"Feature coefficients plot saved for {model} at {feature_coefficients_path}")
    else:
        print(f"No feature importance or coefficients available for model {model}.")

    # Heatmap of Predictions vs True Values
    sns.heatmap(pd.DataFrame({"True": y_test, "Predicted": y_pred}).pivot_table(index='True', columns='Predicted', aggfunc=len, fill_value=0), annot=True, fmt="d", cmap="YlGnBu")
    plt.title("Heatmap of Predictions vs True Values")
    plt.savefig(os.path.join(output_dir, "heatmap_prediction_vs_true.png"))
    plt.close()

    # Scatterplot with Predicted Probabilities
    if y_prob is not None:
        plt.scatter(range(len(y_test)), y_prob, c=y_test, cmap="bwr", alpha=0.6)
        plt.title("Scatterplot of Predicted Probabilities")
        plt.xlabel("Sample Index")
        plt.ylabel("Predicted Probability")
        plt.colorbar(label="True Value")
        plt.savefig(os.path.join(output_dir, "scatterplot_predicted_probabilities.png"))
        plt.close()

    # Metrics
    metrics = {
        'accuracy': accuracy_score(y_test, y_pred),
        'precision': precision_score(y_test, y_pred, zero_division=0),
        'recall': recall_score(y_test, y_pred, zero_division=0),
        'f1_score': f1_score(y_test, y_pred, zero_division=0),
        'roc_auc': roc_auc_score(y_test, y_prob) if y_prob is not None else 'N/A',
        'mse': mse,
        'mae': mae,
        'rmse': rmse,
        'confusion_matrix': cm.tolist(),
        'classification_report': classification_report(y_test, y_pred, zero_division=0)
    }
    return metrics

# Function to list and select models from the model folder
def select_models(model_folder):
    model_files = [f for f in os.listdir(model_folder) if f.endswith('.pkl')]
    if not model_files:
        print("No model files found in the specified model folder.")
        return {}
    
    print("Available models for prediction:")
    for idx, model_file in enumerate(model_files, 1):
        print(f"{idx}. {model_file}")
    
    selection = input("Enter the model numbers to use for prediction (comma-separated) or 'all' to use all models: ")
    if selection.lower() == 'all':
        selected_files = model_files
    else:
        selected_indices = [int(i.strip()) - 1 for i in selection.split(",")]
        selected_files = [model_files[i] for i in selected_indices]
    
    models = {}
    for model_file in selected_files:
        model_path = os.path.join(model_folder, model_file)
        model_name = model_file.split('_model.pkl')[0]
        models[model_name] = joblib.load(model_path)
    
    return models

def main(config_file="config.txt"):
    visialize_output_folder = "../ML_DATA/visualize_output"
    # Load paths from config
    paths = load_paths(config_file)

    # Select and load test file
    test_file = select_testing_file(paths["input_folder_predict"])
    if not test_file:
        print("No valid file selected. Exiting.")
        return

    X_test, y_test = load_data(test_file)

    # Select models to use for prediction
    models = select_models(paths["model_folder_predict"])
    if not models:
        print("No models selected for prediction. Exiting.")
        return

    
    report = []
    for model_name, model in models.items():
        # Extract feature names from the model
        try:
            feature_names = model.feature_names  # Assuming you saved feature_names in training script
        except AttributeError:
            print(f"Could not extract feature names from model {model_name}.")
            continue
        model_output_dir = os.path.join(visialize_output_folder, model_name)
        # metrics = evaluate_model(model, X_test, y_test, model_output_dir)
        metrics = evaluate_model(model, X_test, y_test, model_output_dir, feature_names)
        report.append((model_name, metrics))
        # print(f"Model: {model_name} - Metrics: {metrics}")

        # Format and print the metrics
        print(f"\n{'='*40}")
        print(f"Model: {model_name}")
        print(f"{'-'*40}")
        print("Classification Report:")
        print(metrics['classification_report'])
        print(f"Accuracy       : {metrics['accuracy']:.4f}")
        # if isinstance(metrics['roc_auc'], str):
        #     print(f"ROC AUC        : {metrics['roc_auc']}")
        # else:
        #     print(f"ROC AUC        : {metrics['roc_auc']:.4f}")
        if metrics['roc_auc'] is not None:
            print(f"ROC AUC        : {metrics['roc_auc']:.4f}")
        else:
            print("ROC AUC        : Not applicable (e.g., no probabilities available).")
        print(f"MSE            : {metrics['mse']:.4f}")
        print(f"MAE            : {metrics['mae']:.4f}")
        print(f"RMSE           : {metrics['rmse']:.4f}")
        print("Confusion Matrix:")
        print(metrics['confusion_matrix'])
        print(f"{'='*40}\n")

    report_path = os.path.join(visialize_output_folder, 'prediction_report.csv')
    report_df = pd.DataFrame([{**{'Model': m}, **metrics} for m, metrics in report])
    report_df.to_csv(report_path, index=False)
    print(f"Prediction report with confusion matrices saved to {report_path}")

if __name__ == "__main__":
    # main()
    parser = argparse.ArgumentParser(description="Run model predictions and evaluations.")
    parser.add_argument('--config_file', type=str, default="config.txt", help="Path to the configuration file.")
    
    args = parser.parse_args()
    main(args.config_file)
